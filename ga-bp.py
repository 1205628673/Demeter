import math
import random
import openpyxl
import pickle
import matplotlib.pyplot as plt
import numpy as np
import math
from sklearn.svm import SVR
from sklearn import preprocessing
from sklearn.neural_network import MLPRegressor
from sklearn.model_selection import train_test_split,GridSearchCV
from sklearn.metrics import accuracy_score,r2_score
from sklearn.cross_decomposition import PLSRegression
from sklearn.metrics import mean_squared_error
from sklearn.metrics import mean_absolute_error
from sklearn.model_selection import cross_val_score
from sklearn.model_selection import RandomizedSearchCV
def make_matrix(m, n, fill=0.0):
    mat = []
    for i in range(m):
        mat.append([fill] * n)
    return mat

def sigmoid(x):
    return 1.0 / (1.0 + math.exp(-x))


def sigmoid_derivative(x):
    return x * (1 - x)
class Classification:
    def __init__(self):
        self.featureNum = 0
        self.features_x = None #特征值
        self.features_y = None   #特征标签
        self.initRowColumn()
        self.readFeature()
        self.model = None
    def initRowColumn(self):
        wb = openpyxl.load_workbook('D:\\soil-feature.xlsx')
        sh = wb['Sheet1']
        row = sh.max_row
        column = sh.max_column
        self.featureNum = column - 1
        self.features_x = [[0 for i in range(column - 1)] for j in range(row)]
        self.features_y = [0 for i in range(row)]
    def readFeature(self):
        wb = openpyxl.load_workbook('D:\\soil-feature.xlsx')
        sh = wb['Sheet1']
        for r in range(1,row + 1):
            for c in range(1,self.column):
                data = sh.cell(row = r,column = c)
                self.features_x[r - 1][c - 1] = data.value
            labelIndex = column
            #根据对应som/g^2的值来打上标签
            som = sh.cell(row = r,column = labelIndex).value
            if som > 40:
                self.features_y[r - 1] = 0
            elif som > 30 and som < 40:
                self.features_y[r - 1] = 1
            elif som > 20 and som < 30:
                self.features_y[r - 1] = 2
            elif som > 10 and som < 20:
                self.features_y[r - 1] = 3
            elif som > 6 and som < 10:
                self.features_y[r - 1] = 4
            else:
                self.features_y[r - 1] = 5
        processor = Processing()
        self.features_x, self.features_y = processor.abnormalElimination(self.features_x,self.features_y)
    def reduceDimension(self,individual):
        #根据GA遗传算法的个体来降低维度
        dimension = 0
        selectDimensions = [] #降维后剩下的维度
        #确定剩余维度
        for d in range(self.featureNum):
            if individual >> d & 1 == 1:
                dimension += 1
                selectDimensions.append(d)
        reduceDimensionFeature = [[0 for i in range(dimension)] for j in range(len(self.features_x))]
        #从原特征中取出选中的维度
        for c in range(dimension):
            dim = selectDimensions[c] #选择的维度是哪一维，也就是哪一列
            for r in range(len(self.features_x)):
                reduceDimensionFeature[r][c] = self.features_x[r][dim] #把原来的特征数据选取出来，填到新的特征变量中
        return reduceDimensionFeature
    def train(self,features,label):
        self.model.fit(features,label)
        '''self.model = GridSearchCV(SVR(kernel='linear', gamma=0.1), cv=5,param_grid = { 
            'C': [0.1, 0.5, 1.0],
            'gamma': [0.1, 0.5, 1.0]
        })
        nComponents = len(features[0])
        self.model = 
        self.model.fit(features,label)'''
    def score(self,feature_x_test,feature_y_test):
        accurate = self.model.score(feature_x_test, feature_y_test)
        return accurate
    def predict(self,feature_x_test):
        result = self.model.predict(feature_x_test)
        return result
class Regression(Classification):
    def __init__(self):
        super().__init__()
    def readFeature(self):
        wb = openpyxl.load_workbook('D:\\soil-feature.xlsx')
        sh = wb['Sheet1']
        row = sh.max_row
        column = sh.max_column
        for r in range(1,row + 1):
            for c in range(1,column):
                data = sh.cell(row = r,column = c)
                self.features_x[r - 1][c - 1] = data.value
            labelIndex = column
            self.features_y[r - 1] = sh.cell(row = r,column = labelIndex).value
        processor = Processing()
        self.features_x, self.features_y = processor.abnormalElimination(self.features_x,self.features_y)
        self.features_X = processor.stander(self.features_x) #标准化
class BpnnRegression(Regression):
    def __init__(self):
        super().__init__()
        self.model = MLPRegressor(solver='lbfgs', alpha=1e-5,
                    hidden_layer_sizes=(26, 15), random_state=1, max_iter=10000)
class PlsrRegression(Regression):
    def __init__(self):
        super().__init__()
        self.model = PLSRegression(n_components=len(self.features_x) - 1)
    def train(self,features,labels):
        self.model = PLSRegression(n_components=len(features[0]))
class SvrRegression(Regression):
    def __init__(self):
        super().__init__()
        self.model = SVR(kernel='linear', gamma=0.1)
class PlsrBpnnRegression():
    def __init__(self,plsr = None,bpnn = None):
        self.loadCombinedModel()
    def loadCombinedModel(self):
        with open('bpnn-linear.pickle','rb') as f:
            self.bpnn = pickle.load(f)
        with open('plsr-linear.pickle','rb') as f:
            self.plsr = pickle.load(f)
    def predict(self,plsrFeatures,plsrLabels,bpnnFeatures,bpnnLabels):
        plsrPredict = self.plsr.predict(plsrFeatures).flatten()
        bpnnPredict = self.bpnn.predict(bpnnFeatures).flatten()
        n = len(plsrFeatures)
        sigmaAP = 0
        for i in range(n):
            sigmaAP = sigmaAP + 1 - abs((plsrLabels[i] - plsrPredict[i]) / plsrLabels[i])
        EP = sigmaAP / n
        SP = 0
        for i in range(n):
            SP = SP + (1 - abs((plsrLabels[i] - plsrPredict[i]) / plsrLabels[i]) - EP) ** 2
        sigSP = math.sqrt(SP) / n
        SP = EP * (1 - sigSP)

        n = len(bpnnFeatures)
        sigmaAB = 0
        for i in range(n):
            sigmaAB = sigmaAB + 1 - abs((bpnnLabels[i] - bpnnPredict[i]) / bpnnLabels[i])
        EB = sigmaAB / n
        SB = 0
        for i in range(n):
            SB = SB + (1 - abs((bpnnLabels[i] - bpnnPredict[i]) / bpnnLabels[i]) - EB) ** 2
        sigSB = math.sqrt(SB) / n
        SB = EB * (1 - sigSB)
        k1 = SP / (SP + SB)
        k2 = SB / (SP + SB)
        result = [0 for j in range(len(bpnnFeatures))]
        for i in range(len(result)):
            result[i] = k1 * plsrPredict[i] + k2 * bpnnPredict[i]
        return result
class BpnnClassication(Classification):
    def __init__(self):
        super().__init__()
        self.model = MLPClassifier(solver='lbfgs', alpha=1e-5,
                    hidden_layer_sizes=(26, ), random_state=1)
class SvrClassification(Classification):
    def __init__(self):
        super().__init__()
        self.model =SVR(kernel='linear', gamma=0.1)
class PlsrClassification(Classification):
    def __init__(self):
        super().__init__()
        self.model = PLSRegression(n_components=len(self.column) - 1)
    def train(self,features,labels):
        self.model = PLSRegression(n_components=len(features[0]))
class Processing:
    def normalizer(self,features):
        scaler = preprocessing.MinMaxScaler().fit(features)
        result = scaler.transform(features)
        return result
    def stander(self,features):
        scaler = preprocessing.StandardScaler().fit(features)
        result = scaler.transform(features)
        return result
    def abnormalElimination(self,features,labels):
        #画出每个样本在1000个模型中的占比
        dimensionNum = len(features[0])
        modelRank = []
        appearProbability = [[0 for i in range(1000)] for j in range(len(features))]
        plt.figure(figsize=(10,10))
        for i in range(1000):
            #训练1000个模型
            model = PLSRegression(n_components = dimensionNum)
            x_train,x_test,y_train,y_test = train_test_split(features,labels)
            model.fit(x_train,y_train)
            y_predict = model.predict(x_test)
            PRESS = 0
            for i in range(len(y_predict)):
                PRESS = PRESS + (y_predict[i] - y_test[i]) ** 2 #计算PRESS值 
            rankObj = {
                'model':model,
                'x_train_set':x_train,
                'y_train_set':y_train,
                'press':PRESS[0]
            }
            modelRank.append(rankObj) #建立模型数组
        #将模型数组按PRESS值正序排列
        for i in range(len(modelRank)):
            for j in range(i,len(modelRank)):
                if modelRank[i]['press'] > modelRank[j]['press']:
                    temp = modelRank[i]
                    modelRank[i] = modelRank[j]
                    modelRank[j] = temp
        #记录每个样本在每个模型中是否出现
        plt.subplot(211)
        for i in range(len(labels)):
            for j in range(len(modelRank)):
                if self.searchSample(features[i], modelRank[j]['x_train_set']):
                    appearProbability[i][j] = 1
        for i in range(len(labels)):
            fmn = 0
            modelArr = []
            facArr = []
            for j in range(len(modelRank)):
                #累加Fmn值
                fmn = fmn + appearProbability[i][j]
                #根据每个样本在当前累计模型的Fac值
                fac = 100 * (fmn / (j + 1)) 
                modelArr.append(j)
                facArr.append(fac)
            #画图
            plt.plot(modelArr, facArr,color = 'black', label = 'Appear probability',LineWidth = 0.5)
        plt.title('esample appear probability in model')
        plt.xlabel('model')
        plt.ylabel('fac')
        #画出每个样本在累计到第50个样本时的Fac值
        fmn = 0
        plt.subplot(212)
        sampleArr = []
        facArr = []
        for i in range(len(labels)):
            fmn = 0
            for j in range(25):
                fmn = fmn + appearProbability[i][j]
            fac = 100 * fmn / (j + 1)
            facArr.append(fac)
            sampleArr.append(i)
        plt.plot(sampleArr, facArr, LineWidth = 0.5)
        plt.xlabel('sample number')
        plt.ylabel('fac')
        for i in range(len(facArr)):
            if facArr[i] >= 100:
                #标出与其他样本有显著差异的异常样本
                plt.text(i,facArr[i],str(i),wrap = True)
                #移除异常样本
                del features[i]
                del labels[i]
        plt.show()
        return features,labels

    def searchSample(self, x, x_train):
        for sample in x_train:
            sfirstValue = sample[0] #模型训练集的首值
            slastValue = sample[len(sample) - 1] #模型训练集的尾值
            xfirstValue = x[0] #要查找的样本的首值
            xlastValue = x[len(x) - 1] #要查找的样本的尾值
            if sfirstValue == xfirstValue and slastValue == xlastValue:
                return True
        return False                
class GA:
    def __init__(self):
        self.population = [] #种群
        self.individualFit = [] #每一轮的个体拟合度
        self.generation = 100 #代数
        self.populationNum = 100 #种群个体数
        self.crossConstant = 0.7 #交叉几率
        self.mutationConstant = 0.2 #变异几率
        self.satisfied = 0.88 #适应度满意值
        self.cls = None #分类器初始化
        self.modelFile = 'plsr.pickle'
        self.xlsFile = 'plsr-dimension-reduce.xlsx'
        self.individualFile = 'plsr-individual.txt'
        for i in range(self.populationNum):
            self.generatePoplation()
    def generatePoplation(self):
        individual = 1
        for bit in range(39):
            if random.random() > 0.5:
                individual = individual << 1
            else:
                individual = (individual << 1) | 1
        self.population.append(individual)
    def select(self):
        #轮盘赌选择方法,通过适应度对比，把适应度低的个体排除
        subPopulation = []
        sumAllFitness = sum(self.individualFit) #所有个体适应度总和
        for i in range(len(self.individualFit)):
            accurate = self.individualFit[i]
            individual = self.population[i]
            present = sumAllFitness / len(self.individualFit)
            if accurate > present:
                #如果该个体的适应度高于平均适应度，则保留该个体,否则淘汰
                subPopulation.append(individual)
        self.population = subPopulation
        #如果种群数量少于规定的一般，则将种群中一半的个体复制一遍
        while len(self.population) < self.populationNum / 2:
            bound = len(self.population) // 2
            self.population.extend(self.population[:bound])
        self.individualFit = [] #清空个体适应度列表
        return present #返回该次选择个体函数的平均适应度
    def fitness(self,individual):
        reduceDimensionFeature = self.cls.reduceDimension(individual)
        feature_y = self.cls.features_y
        x_train,x_test,y_train,y_test = train_test_split(reduceDimensionFeature,self.cls.features_y)
        self.cls.train(x_train,y_train) #调用一下train函数，初始化plsr模型
        self.cls.model.fit(x_train,y_train)
        #交叉验证
        scores = cross_val_score(self.cls.model, x_train, y_train, cv = 5)
        #适应函数用SVR,PLSR等智能算法的准确度
        return scores.mean()
    def crossSupport(self,individual,index,newBit):
        if newBit > 1:
            raise Exception('invalid new bit')
        if individual >> index & 1 == 0:
            individual |= newBit << index 
        else:
            individual &= ~(newBit << index)
        return individual
    def cross(self, theIndex, anotherIndex):
        #如果n不是最后一个个体，就和下一个个体做交叉。
        theIndividual = self.population[theIndex]
        anotherIndividual = self.population[anotherIndex]
        for crossCount in range(5): #交叉五个点位
            crossPosition = int(random.random()*40) #随机在维度范围内选出要交叉的点位
            #获取两个个体要交叉的bit位值
            theBit = theIndividual >> crossPosition & 1
            anotherBit = anotherIndividual >> crossPosition & 1
            #做bit位交叉
            theIndividual = self.crossSupport(theIndividual,crossPosition,anotherBit)
            anotherIndividual = self.crossSupport(anotherIndividual,crossPosition,theBit)
        #将交叉得出的新个体加入种群
        self.population[theIndex] = theIndividual 
        self.population[anotherIndex] = anotherIndividual

    def mutation(self,x):
        if random.random() > 0.5:
            mutationPosition = int(random.random()*40)
            if x >> mutationPosition & 1 == 0:
                x |= 1 << mutationPosition
            else:
                x &= ~(1 << mutationPosition)
        return x
    def evolution(self):
        roundCount = 0      #进化轮次
        bestFitness = 0.0   #最佳适应度
        bestIndividual = 0       #最佳适应度个体
        bestFeatures = []   #最佳适应度的特征
        #nn = BPNeuralNetwork()
        presentArr = [] #用于画出每轮选择个体函数的平均适应度的数组
        bestFitnessArr = [] #用于画出每轮选择的最佳适应度的数组
        while bestFitness < self.satisfied and roundCount < self.generation:
            bestFitnessArr.append(bestFitness) #将本次最好的适应度加入数组，用于画图
            #交叉
            i = 0
            while i < len(self.population) - 1:
                if(random.random() > self.crossConstant and self.population[i] != bestIndividual):
                    self.cross(i, i + 1)
                i = i + 1
            #变异
            i = 0
            while i < len(self.population) - 1:
                if(random.random() > self.mutationConstant and self.population[i] != bestIndividual):
                    mutationIndividual = self.mutation(self.population[i]) #得到变异个体
                    self.population[i] = mutationIndividual #将种群中的该个体修改为变异出来的个体
                i = i + 1
            #计算所有个体的适应度
            for i in self.population:
                accurate = self.fitness(i)
                print('fit is %s,and individual %s' %(accurate,bin(i)))
                self.individualFit.append(accurate)
            #获得个体中的最优适应度
                if accurate > bestFitness:
                    bestFitness = accurate
                    bestIndividual = i
                    bestFeatures = self.cls.reduceDimension(i)
                    if bestFitness >= self.satisfied:
                        bestFitnessArr.append(bestFitness)
                        break
            #选择子代
            present = self.select()
            presentArr.append(present)
            roundCount = roundCount + 1
        plt.plot([i for i in range(len(presentArr))], presentArr)
        plt.plot([i for i in range(len(bestFitnessArr))], bestFitnessArr)
        plt.xlabel('Generate Round')
        plt.ylabel('Fitness Value')
        plt.show()
        #如果进化轮次大于设置，则退出方法
        if(roundCount > self.generation):
            return
        self.fitness(bestIndividual) #找到最佳的个体后进行Fit，建立出模型
        #保存训练好的模型
        with open(self.modelFile,'wb') as f:
            pickle.dump(self.cls.model,f)
            print('模型已保存')
        #根据有最佳适应度的个体保存对应降维后的数据
        features = self.cls.reduceDimension(bestIndividual)
        labels = self.cls.features_y
        dimensionNum = len(features[0]) #获取维度数
        data = [[0 for i in range(dimensionNum + 1)] for j in range(len(features))]
        #保存标签
        for i in range(len(labels) - 1):
            data[i][dimensionNum] = labels[i]
        #保存降维后的电压值
        for r in range(len(features)):
            for c in range(len(features[0])):
                data[r][c] = features[r][c]
        print(data)
        try:
            wb = openpyxl.Workbook()
            workObject = wb.active
            sheet = wb.create_sheet('Sheet1')
            for row in data:
                sheet.append(row)
            wb.save(self.xlsFile)
            print('已保存降维的电压值')
            f = open(self.individualFile,'w')
            f.write(str(bin(bestIndividual)))
            f.close()
        except Exception as error:
            print('happen some error in save excel file...,%s' % error)
class Draw:
    def __init__(self,modelFile,individualFile):
        self.model = None
        self.features = None
        self.featuresNum = 0
        self.featuresIndex = None
        self.label = None
        self.modelFile = modelFile
        self.individualFile = individualFile
        self.loadModel()
        self.features, self.label = self.loadFeature(individualFile)
    def loadModel(self):
        with open(self.modelFile,'rb') as f:
            self.model = pickle.load(f)
    def loadFeature(self,individualFile):
        #根据个体二进制的表现型来读取选择的维度
        featureNum = 0 #记录维度数
        featureIndexs = [] #维度下标
        try:
            f = open(individualFile)
            value = f.read()
        except Exception as e:
            print(e)
        individual = int(value,2) #二进制串转成数字
        for i in range(40):
            if individual >> i & 1 == 1:
                featureNum = featureNum + 1
                featureIndexs.append(i)
        self.featuresNum = featureNum
        self.featuresIndex = featureIndexs
        wb = openpyxl.load_workbook('D:\\soil-feature.xlsx')
        sh = wb['Sheet1']
        row = sh.max_row
        column = sh.max_column
        features = [[0 for i in range(featureNum)] for j in range(row)]
        label = [0 for i in range(row)]
        #将选择的维度的电压值载入
        for c in range(featureNum):
            dim = featureIndexs[c]
            for r in range(1,row + 1):
                data = sh.cell(row = r, column = dim + 1)
                features[r - 1][c] = data.value
        #载入对应的标签
        for r in range(1,row + 1):
            labelIndex = column
            som = sh.cell(row = r,column = labelIndex).value
            label[r - 1] = som
        return features, label
    def bpnnTest(self):
        #寻找bpnn最优隐含层节点数
        param = dict(hidden_layer_sizes=range(1,100))
        x_train,x_test,y_train,y_test = train_test_split(self.features,self.label)
        grid = RandomizedSearchCV(estimator=MLPRegressor(), param_distributions=param)
        grid_result = grid.fit(x_train, y_train)
        print('Best：%f using %s' % (grid_result.best_score_, grid_result.best_params_))
        '''
        RMSEArr = []
        RMSERArr = []
        nodesMax = 100#self.featuresNum * 2 // 3
        for n in range(1,nodesMax):
            self.model = MLPRegressor(solver='lbfgs', alpha=1e-5,
                    hidden_layer_sizes=(n, ), random_state=1, max_iter=10000)
            x_train,x_test,y_train,y_test = train_test_split(self.features,self.label)
            self.model.fit(x_train,y_train)
            preds = self.model.predict(x_test)
            predsR = self.model.predict(x_train)
            preds = preds.flatten()
            sigma = 0
            sigmaR = 0
            for i in range(len(x_test)):
                sigma = sigma + (preds[i] - y_test[i]) ** 2
            for i in range(len(x_train)):
                sigmaR = sigmaR + (predsR[i] - y_train[i]) ** 2
            MSE = sigma / len(x_test)
            RMSE = math.sqrt(MSE)
            RMSER = math.sqrt(sigmaR / len(x_train))
            RMSEArr.append(RMSE)
            RMSERArr.append(RMSER)
        generalizerError, = plt.plot([i for i in range(1,nodesMax)], RMSEArr)
        trainError, = plt.plot([i for i in range(1,nodesMax)], RMSERArr,color = 'r')
        plt.xlabel('number of hidden layer nodes')
        plt.ylabel('RMSE')
        plt.legend([generalizerError, trainError], ['generalizer error','train error'], loc='upper left')
        plt.show()'''
    def drawCombinedPredict(self):
        bpnnFeatures,bpnnLabels = self.loadFeature('bpnn-individual.txt')
        plsrFeatures,plsrLabels = self.loadFeature('plsr-individual.txt')
        plsrBpnn = PlsrBpnnRegression()
        preds = plsrBpnn.predict(plsrFeatures, plsrLabels, bpnnFeatures, bpnnLabels)
        rmse = mean_squared_error(bpnnLabels,preds) ** 0.5
        mase = mean_absolute_error(bpnnLabels,preds)
        self.loadModel()
        bpnnPreds = self.model.predict(bpnnFeatures)
        plt.figure(figsize=(10,7))
        plt.plot([i for i in range(len(bpnnLabels))], bpnnPreds, 'b--o')
        plt.plot([i for i in range(len(bpnnLabels))], preds,'r--*')
        plt.title('R^2=%f\nRMSE=%f\nMASE=%f'%(r2_score(bpnnLabels, preds),rmse,mase))
        plt.xlabel('sample number')
        plt.ylabel('predict value')
        plt.show()
    def drawFit(self):
        x_train,x_test,y_train,y_test = train_test_split(self.features,self.label)
        x_test = self.features
        y_test = self.label
        preds = self.model.predict(x_test)
        rmse = mean_squared_error(y_test,preds) ** 0.5
        mase = mean_absolute_error(y_test,preds)
        sig = 0
        #计算平均误差率MAPE
        for i in range(len(preds)):
            sig = sig + abs((y_test[i] - preds[i]) / y_test[i])
        mape = sig / len(preds) * 100
        p = []
        parameter = np.polyfit(preds.flatten(), np.array(y_test), deg=2)
        y = np.polyval(parameter, preds)
        #画线图
        plt.figure(figsize=(10,7))
        obs, = plt.plot([i for i in range(len(y_test))], y_test, 'b--o')
        p, = plt.plot([i for i in range(len(y_test))], preds,'r--*')
        plt.xlabel('sample number')
        plt.ylabel('predict value')
        plt.legend([obs, p], ['observe', 'predict'])
        plt.title('R^2=%f\nRMSE=%f\nMAPE=%f'%(r2_score(y_test, preds),rmse,mape))
        plt.show()
        #画点图
        plt.figure(figsize=(10,7))
        plt.ylabel('observe value')
        plt.xlabel('predict value')
        plt.title('R^2=%f\nRMSE=%f\nMAPE=%f'%(r2_score(y_test, preds),rmse,mape))
        plt.plot(preds, y, color='g')
        point = plt.scatter(preds, y_test, c='#00CED1', alpha=0.4, label='预测点')
        plt.legend([point],['sample value'])
        plt.show()
    def drawBox(self):
        #画箱形图
        cls = Classification()
        somfeature = cls.features_x #xls表中有机质的电压特征
        boxData = [[0 for i in range(len(somfeature))] for j in range(len(somfeature[0]))]
        boxLabels = []
        for r in range(len(somfeature)):
            for c in range(len(somfeature[0])):
                boxData[c][r] = somfeature[r][c]
        for i in range(len(somfeature[0])):
            if i // 10 == 0:
                boxLabels.append('S%d-max'%(i%10 + 1))
            elif i // 10 == 1:
                boxLabels.append('S%d-MDCV'%(i%10 + 1))
            elif i // 10 == 2:
                boxLabels.append('S%d-RAV'%(i%10 + 1))
            else:
                boxLabels.append('S%d-Vt'%(i%10 + 1))
        plt.figure(figsize=(10,7))
        plt.xticks(rotation=90)
        plt.boxplot(boxData,labels = boxLabels,widths = 0.5)
        plt.show()
if __name__ == '__main__':
    '''
    ga = GA()
    ga.modelFile = 'svr-linear.pickle'
    ga.xlsFile = 'svr-dimension-reduce.xlsx'
    ga.individualFile = 'svr-individual.txt'
    ga.cls = SvrRegression()
    ga.evolution()
    '''
    bpnnDraw = Draw('bpnn-linear.pickle', 'bpnn-individual.txt')
    bpnnDraw.drawFit()
    svrDraw = Draw('svr-linear.pickle', 'svr-individual.txt')
    svrDraw.drawFit()
    plsrDraw = Draw('plsr-linear.pickle', 'plsr-individual.txt')
    plsrDraw.drawFit()
    bpnnDraw.model = PlsrBpnnRegression()
    bpnnDraw.drawCombinedPredict()
